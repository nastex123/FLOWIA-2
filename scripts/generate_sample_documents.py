"""Script to generate realistic business sample documents (XLSX, CSV, PDF) for FlowMind AI testing."""

from pathlib import Path
import csv
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import fitz  # PyMuPDF


def ensure_samples_dir() -> Path:
    """Creates samples directory in repository root."""
    samples_dir = Path(__file__).parent.parent / "samples"
    samples_dir.mkdir(parents=True, exist_ok=True)
    return samples_dir


def generate_invoice_xlsx(output_dir: Path) -> Path:
    """Generates a professional business invoice spreadsheet."""
    file_path = output_dir / "factura_suministros_2024.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Factura F-2024-0982"

    # Styling definitions
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    title_font = Font(name="Calibri", size=16, bold=True, color="0F172A")
    meta_bold = Font(name="Calibri", size=10, bold=True)
    meta_regular = Font(name="Calibri", size=10)
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    # Title & Vendor
    ws["A1"] = "SUMINISTROS INDUSTRIALES IBERIA S.L."
    ws["A1"].font = title_font
    ws["A2"] = "CIF: B-87654321 | Email: facturacion@suministros-iberia.com"
    ws["A2"].font = meta_regular
    ws["A3"] = "Polígono Industrial Las Eras, Nave 4, Madrid, España"
    ws["A3"].font = meta_regular

    # Metadata Block
    ws["E1"] = "Nº FACTURA:"
    ws["E1"].font = meta_bold
    ws["F1"] = "F-2024-0982"
    ws["F1"].font = Font(name="Calibri", size=11, bold=True, color="4F46E5")

    ws["E2"] = "FECHA EMISIÓN:"
    ws["E2"].font = meta_bold
    ws["F2"] = "2024-06-15"
    ws["F2"].font = meta_regular

    ws["E3"] = "FECHA VENCIMIENTO:"
    ws["E3"].font = meta_bold
    ws["F3"] = "2024-07-15"
    ws["F3"].font = meta_regular

    # Client Block
    ws["A5"] = "DATOS DEL CLIENTE"
    ws["A5"].font = meta_bold
    ws["A6"] = "Razón Social: Construcciones y Reformas del Norte S.A."
    ws["A7"] = "NIF / CIF: A-11223344"
    ws["A8"] = "Dirección: Avda. Principal 45, Bilbao, Vizcaya"

    # Line Items Header
    headers = ["Referencia", "Descripción de Producto / Servicio", "Cantidad", "Precio Unitario (€)", "Importe (€)"]
    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=10, column=col_num)
        cell.value = header_title
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center" if col_num in (1, 3) else "left")

    # Line Items Data
    items = [
        ("REF-8801", "Servidor Rack 2U Dell PowerEdge R750", 2, 2450.00, 4900.00),
        ("REF-8802", "Switch Gestionable Cisco Catalyst 48 Puertos PoE", 4, 850.50, 3402.00),
        ("REF-8803", "Módulo Transceptor SFP+ 10Gbps LC", 10, 45.00, 450.00),
        ("REF-8804", "Bobina Cable UTP Cat6A 305m Libre Halógenos", 5, 120.00, 600.00),
        ("REF-8805", "Instalación, Certificación y Puesta en Marcha", 16, 65.00, 1040.00),
    ]

    for row_idx, item in enumerate(items, 11):
        for col_idx, val in enumerate(item, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = val
            cell.border = thin_border
            if col_idx in (4, 5):
                cell.number_format = "#,##0.00 €"
                cell.alignment = Alignment(horizontal="right")
            elif col_idx == 3:
                cell.alignment = Alignment(horizontal="center")

    # Totals Block
    ws["D17"] = "Base Imponible:"
    ws["D17"].font = meta_bold
    ws["E17"] = 10392.00
    ws["E17"].number_format = "#,##0.00 €"

    ws["D18"] = "IVA (21%):"
    ws["D18"].font = meta_bold
    ws["E18"] = 2182.32
    ws["E18"].number_format = "#,##0.00 €"

    ws["D19"] = "TOTAL A PAGAR:"
    ws["D19"].font = Font(name="Calibri", size=12, bold=True, color="0F172A")
    ws["E19"] = 12574.32
    ws["E19"].font = Font(name="Calibri", size=12, bold=True, color="4F46E5")
    ws["E19"].number_format = "#,##0.00 €"

    # Payment info
    ws["A21"] = "Forma de Pago: Transferencia Bancaria a 30 días"
    ws["A21"].font = meta_bold
    ws["A22"] = "IBAN: ES7621000418401234567891 | BIC: CAIXESBBXXX"

    # Adjust Column Widths
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 48
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 22
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 18

    wb.save(file_path)
    return file_path


def generate_inventory_csv(output_dir: Path) -> Path:
    """Generates an inventory / stock list CSV with European delimiter."""
    file_path = output_dir / "inventario_almacen_central.csv"
    data = [
        ["SKU", "Descripcion_Articulo", "Categoria", "Stock_Actual", "Stock_Minimo", "Coste_Unitario", "Ubicacion_Pasillo"],
        ["HW-SRV-01", "Servidor ProLiant DL380 Gen10", "Servidores", 14, 5, 2100.50, "P01-E03"],
        ["HW-SWT-02", "Switch Ethernet 24 Puertos Gigabit", "Networking", 32, 10, 420.00, "P02-E01"],
        ["HW-RTR-03", "Router VPN Empresarial Dual-WAN", "Networking", 8, 4, 680.75, "P02-E04"],
        ["ACC-CAB-04", "Latiguillo RJ45 Cat6A 2m Gris", "Accesorios", 250, 50, 3.20, "P05-E01"],
        ["ACC-CAB-05", "Latiguillo Fibra Optica LC-LC 5m", "Accesorios", 85, 20, 12.50, "P05-E02"],
        ["PWR-UPS-06", "Sistema SAI Online 3000VA Rack", "Energia", 6, 2, 1150.00, "P03-E02"],
        ["STO-HDD-07", "Disco Duro Enterprise SATA 8TB 7200RPM", "Almacenamiento", 45, 15, 185.00, "P04-E01"],
        ["STO-SSD-08", "Unidad SSD NVMe PCIe 4.0 2TB", "Almacenamiento", 60, 20, 140.00, "P04-E03"],
    ]

    with open(file_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f, delimiter=";")
        writer.writerows(data)

    return file_path


def generate_purchase_order_csv(output_dir: Path) -> Path:
    """Generates a purchase order CSV document."""
    file_path = output_dir / "orden_compra_material_po_4091.csv"
    data = [
        ["PO_Number", "Order_Date", "Vendor_Name", "Vendor_Tax_ID", "Item_Code", "Description", "Quantity", "Unit_Price", "Total_Price"],
        ["PO-4091", "2024-06-10", "TechSupply Distribution SL", "B-99887766", "ITM-101", "Monitor Profesional 27 IPS 4K", 10, 299.00, 2990.00],
        ["PO-4091", "2024-06-10", "TechSupply Distribution SL", "B-99887766", "ITM-102", "Teclado Mecanico Inalambrico", 10, 75.00, 750.00],
        ["PO-4091", "2024-06-10", "TechSupply Distribution SL", "B-99887766", "ITM-103", "Raton Ergonomico Bluetooth", 10, 45.50, 455.00],
        ["PO-4091", "2024-06-10", "TechSupply Distribution SL", "B-99887766", "ITM-104", "Docking Station USB-C Triple Display", 10, 160.00, 1600.00],
    ]

    with open(file_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f, delimiter=",")
        writer.writerows(data)

    return file_path


def generate_payroll_xlsx(output_dir: Path) -> Path:
    """Generates a payroll / payslip spreadsheet."""
    file_path = output_dir / "nomina_empleado_junio_2024.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Nómina Junio 2024"

    ws["A1"] = "RECIBO INDIVIDUAL JUSTIFICATIVO DEL PAGO DE SALARIOS"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="1E293B")

    ws["A3"] = "Empresa: Soluciones Digitales FlowMind S.L."
    ws["A4"] = "CIF: B-44556677"
    ws["A5"] = "Código Cuenta Cotización: 28/1234567/89"

    ws["E3"] = "Trabajador: Alejandro Martínez Gómez"
    ws["E4"] = "NIF: 47891234X"
    ws["E5"] = "Categoría: Ingeniero de Software Senior"

    headers = ["Concepto Salarial", "Devengos (€)", "Deducciones (€)"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=7, column=col)
        cell.value = h
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="334155", end_color="334155", fill_type="solid")

    rows = [
        ("Salario Base", 2800.00, None),
        ("Plus Convenio", 450.00, None),
        ("Complemento de Puesto", 300.00, None),
        ("Aportación Seguridad Social (4.7%)", None, 166.85),
        ("Desempleo (1.55%)", None, 55.03),
        ("Formación Profesional (0.1%)", None, 3.55),
        ("Retención IRPF (17%)", None, 603.50),
    ]

    for r_idx, (concept, devengo, deduc) in enumerate(rows, 8):
        ws.cell(row=r_idx, column=1, value=concept)
        c2 = ws.cell(row=r_idx, column=2, value=devengo)
        c3 = ws.cell(row=r_idx, column=3, value=deduc)
        if devengo:
            c2.number_format = "#,##0.00 €"
        if deduc:
            c3.number_format = "#,##0.00 €"

    ws["A16"] = "Total Devengado (Bruto):"
    ws["A16"].font = Font(bold=True)
    ws["B16"] = 3550.00
    ws["B16"].number_format = "#,##0.00 €"

    ws["A17"] = "Total Deducciones:"
    ws["A17"].font = Font(bold=True)
    ws["C17"] = 828.93
    ws["C17"].number_format = "#,##0.00 €"

    ws["A19"] = "LÍQUIDO TOTAL A PERCIBIR (NETO):"
    ws["A19"].font = Font(name="Calibri", size=12, bold=True, color="4F46E5")
    ws["B19"] = 2721.07
    ws["B19"].font = Font(name="Calibri", size=12, bold=True, color="4F46E5")
    ws["B19"].number_format = "#,##0.00 €"

    ws.column_dimensions['A'].width = 38
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['E'].width = 40

    wb.save(file_path)
    return file_path


def generate_invoice_pdf(output_dir: Path) -> Path:
    """Generates a high-quality PDF invoice with PyMuPDF."""
    file_path = output_dir / "factura_consultoria_cloud.pdf"
    doc = fitz.open()
    page = doc.new_page(width=595, height=842)  # A4

    # Header title
    page.insert_text((50, 60), "CLOUD ARCHITECTURE & DATA SOLUTIONS S.L.", fontsize=14, fontname="helv", color=(0.1, 0.15, 0.3))
    page.insert_text((50, 78), "CIF: B98765432 | Email: billing@cloudarch-solutions.com", fontsize=9, fontname="helv", color=(0.3, 0.3, 0.3))
    page.insert_text((50, 92), "Paseo de la Castellana 120, Madrid, España", fontsize=9, fontname="helv", color=(0.3, 0.3, 0.3))

    # Invoice Meta (Right-aligned)
    page.insert_text((380, 60), "FACTURA: FAC-2024-0718", fontsize=12, fontname="helv", color=(0.2, 0.2, 0.8))
    page.insert_text((380, 78), "Fecha: 2024-06-18", fontsize=9, fontname="helv", color=(0.3, 0.3, 0.3))
    page.insert_text((380, 92), "Vencimiento: 2024-07-18", fontsize=9, fontname="helv", color=(0.3, 0.3, 0.3))

    # Line Separator
    page.draw_line((50, 110), (545, 110), color=(0.7, 0.7, 0.7), width=1)

    # Client Information
    page.insert_text((50, 135), "FACTURAR A:", fontsize=10, fontname="helv", color=(0.2, 0.2, 0.2))
    page.insert_text((50, 150), "Empresa Cliente: Logística Global del Mediterráneo S.A.", fontsize=10, fontname="helv")
    page.insert_text((50, 165), "NIF / CIF: A76543219", fontsize=10, fontname="helv")
    page.insert_text((50, 180), "Dirección: Puerto de Valencia, Muelle Este 12, Valencia", fontsize=9, fontname="helv", color=(0.3, 0.3, 0.3))

    # Table Header Box
    page.draw_rect(fitz.Rect(50, 210, 545, 230), color=(0.15, 0.2, 0.35), fill=(0.15, 0.2, 0.35))
    page.insert_text((60, 224), "Descripción del Servicio", fontsize=9, fontname="helv", color=(1, 1, 1))
    page.insert_text((330, 224), "Horas", fontsize=9, fontname="helv", color=(1, 1, 1))
    page.insert_text((400, 224), "Tarifa / Hora", fontsize=9, fontname="helv", color=(1, 1, 1))
    page.insert_text((490, 224), "Subtotal", fontsize=9, fontname="helv", color=(1, 1, 1))

    # Table Content
    items = [
        ("Auditoría de Arquitectura Cloud & Seguridad", "30", "90.00 €", "2,700.00 €"),
        ("Migración de Base de Datos PostgreSQL a On-Premise", "45", "85.00 €", "3,825.00 €"),
        ("Optimización de Pipelines de Ingesta Tabular", "25", "80.00 €", "2,000.00 €"),
        ("Configuración de Monitorización y Telemetría", "15", "75.00 €", "1,125.00 €"),
    ]

    y = 250
    for desc, hrs, rate, sub in items:
        page.insert_text((60, y), desc, fontsize=9, fontname="helv", color=(0.1, 0.1, 0.1))
        page.insert_text((340, y), hrs, fontsize=9, fontname="helv", color=(0.1, 0.1, 0.1))
        page.insert_text((405, y), rate, fontsize=9, fontname="helv", color=(0.1, 0.1, 0.1))
        page.insert_text((485, y), sub, fontsize=9, fontname="helv", color=(0.1, 0.1, 0.1))
        page.draw_line((50, y + 6), (545, y + 6), color=(0.85, 0.85, 0.85), width=0.5)
        y += 24

    # Summary Totals Box
    y_totals = y + 20
    page.insert_text((370, y_totals), "Base Imponible:", fontsize=10, fontname="helv", color=(0.2, 0.2, 0.2))
    page.insert_text((480, y_totals), "9,650.00 €", fontsize=10, fontname="helv", color=(0.1, 0.1, 0.1))

    page.insert_text((370, y_totals + 18), "IVA (21%):", fontsize=10, fontname="helv", color=(0.2, 0.2, 0.2))
    page.insert_text((480, y_totals + 18), "2,026.50 €", fontsize=10, fontname="helv", color=(0.1, 0.1, 0.1))

    page.draw_line((370, y_totals + 28), (545, y_totals + 28), color=(0.5, 0.5, 0.5), width=1)

    page.insert_text((370, y_totals + 44), "IMPORTE TOTAL:", fontsize=11, fontname="helv", color=(0.1, 0.1, 0.4))
    page.insert_text((475, y_totals + 44), "11,676.50 €", fontsize=12, fontname="helv", color=(0.2, 0.2, 0.8))

    # Footer Payment Info
    page.insert_text((50, 750), "Datos de Pago y Transferencia:", fontsize=9, fontname="helv", color=(0.2, 0.2, 0.2))
    page.insert_text((50, 765), "Titular: CLOUD ARCHITECTURE & DATA SOLUTIONS S.L.", fontsize=8, fontname="helv", color=(0.4, 0.4, 0.4))
    page.insert_text((50, 778), "IBAN: ES9121000418451234567890 | SWIFT/BIC: BSCHESMMXXX", fontsize=8, fontname="helv", color=(0.4, 0.4, 0.4))

    doc.save(file_path)
    doc.close()
    return file_path


def generate_contract_pdf(output_dir: Path) -> Path:
    """Generates a business contract agreement PDF."""
    file_path = output_dir / "contrato_prestacion_servicios.pdf"
    doc = fitz.open()
    page = doc.new_page(width=595, height=842)

    page.insert_text((50, 60), "CONTRATO DE PRESTACIÓN DE SERVICIOS MERCANTILES", fontsize=13, fontname="helv", color=(0.1, 0.1, 0.3))
    page.insert_text((50, 80), "Referencia Contrato: CNT-2024-0088", fontsize=10, fontname="helv", color=(0.4, 0.4, 0.4))

    body_text = """
    REUNIDOS:
    
    De una parte, FLOWMIND AUTOMATION SYSTEMS S.L., con CIF B-12345678, y domicilio en Paseo de la Castellana 200, Madrid, representada por D. Carlos Ruiz en calidad de Director General.
    
    De otra parte, INNOVACIÓN Y TECNOLOGÍA IBÉRICA S.A., con NIF A-87654321, y domicilio en Avda. Diagonal 640, Barcelona, representada por Dña. Laura Sánchez en calidad de Apoderada.
    
    Ambas partes se reconocen mutuamente capacidad jurídica suficiente para formalizar el presente contrato y:
    
    ACUERDAN Y ESTIPULAN:
    
    PRIMERA. — OBJETO DEL CONTRATO: La empresa proveedora prestará servicios de análisis, procesamiento automatizado y clasificación de documentos mediante infraestructura local de Machine Learning.
    
    SEGUNDA. — CONFIDENCIALIDAD Y PRIVACIDAD: Las partes acuerdan que todos los datos procesados permanecerán en servidores locales sin transferencia a APIs de terceros (Zero Cloud Data Leakage).
    
    TERCERA. — PRECIO Y FORMA DE PAGO: El importe total acordado es de 4,500.00 € mensuales, pagaderos mediante domiciliación bancaria dentro de los primeros 5 días hábiles de cada mes.
    
    CUARTA. — JURISDICCIÓN: Para cualquier controversia, las partes se someten a los juzgados y tribunales de la ciudad de Madrid.
    """

    y = 110
    for line in body_text.strip().split("\n"):
        page.insert_text((50, y), line.strip(), fontsize=9, fontname="helv", color=(0.2, 0.2, 0.2))
        y += 15

    # Signatures
    page.insert_text((50, 680), "POR EL PROVEEDOR:", fontsize=9, fontname="helv", color=(0.2, 0.2, 0.2))
    page.insert_text((50, 695), "D. Carlos Ruiz (Director General)", fontsize=8, fontname="helv", color=(0.4, 0.4, 0.4))

    page.insert_text((350, 680), "POR EL CLIENTE:", fontsize=9, fontname="helv", color=(0.2, 0.2, 0.2))
    page.insert_text((350, 695), "Dña. Laura Sánchez (Apoderada)", fontsize=8, fontname="helv", color=(0.4, 0.4, 0.4))

    doc.save(file_path)
    doc.close()
    return file_path


def main():
    print("=" * 60)
    print("   Generando archivos de prueba para FlowMind AI...")
    print("=" * 60)

    out_dir = ensure_samples_dir()

    f1 = generate_invoice_xlsx(out_dir)
    print(f" [OK] Factura Excel generada:    {f1.name}")

    f2 = generate_inventory_csv(out_dir)
    print(f" [OK] Inventario CSV generado:   {f2.name}")

    f3 = generate_purchase_order_csv(out_dir)
    print(f" [OK] Orden de Compra CSV:       {f3.name}")

    f4 = generate_payroll_xlsx(out_dir)
    print(f" [OK] Nómina Salarial Excel:     {f4.name}")

    f5 = generate_invoice_pdf(out_dir)
    print(f" [OK] Factura en PDF generada:   {f5.name}")

    f6 = generate_contract_pdf(out_dir)
    print(f" [OK] Contrato Mercantil PDF:    {f6.name}")

    print("\n -> Todos los archivos se han guardado exitosamente en:")
    print(f"    {out_dir.resolve()}")
    print("=" * 60)


if __name__ == "__main__":
    main()
